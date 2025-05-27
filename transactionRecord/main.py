# main.py

from src.csv_parser import parse_zfb_csv
from src.csv_parser import parse_wx_csv
import shutil
import zipfile
import os
import util
import openpyxl  # 导入 openpyxl 库，用于处理 Excel 文件
from rich.console import Console
from datetime import datetime
from rich.table import Table
from collections import defaultdict

def main():
    bill_path = os.path.expanduser("~") + '/Downloads/账单/202504/'
    name_with_passwd = {
        "Me-微信.zip": "293464", 
        "Me-支付宝.zip": "903386", 
        "Wife-微信.zip": "916923",
        "Wife-支付宝.zip": "634073"
    }

    for file_name, passwd in name_with_passwd.items():
        unzip_file(bill_path + file_name, passwd)
 
    wx_tx_total = handle_wx_excel(bill_path)
    zfb__tx_total = handle_zfb_excel(bill_path)
    xlsx_export(bill_path + "账单合并.xlsx", wx_tx_total + zfb__tx_total)

def unzip_file(zip_file_path, password):
    # 获取解压目标路径和文件名
    last_slash_index = zip_file_path.rfind('/')
    unzip_target_path = zip_file_path[:last_slash_index + 1]
    dot_zip_index = zip_file_path.rfind('.zip')
    file_name = zip_file_path[last_slash_index + 1:dot_zip_index]

    try:
        # 打开 zip 文件
        with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
            # 设置解压密码
            zip_ref.setpassword(password.encode('utf-8'))      

            temp_file = None
            temp_path = None
            for contain_file_name in zip_ref.namelist():
                if not contain_file_name.endswith('.csv'):
                    continue
                # 将 csv 文件解压到指定目录
                temp_file = contain_file_name
                temp_path = temp_file[:temp_file.rfind('/') + 1]
                zip_ref.extract(contain_file_name, unzip_target_path)
            
            # 移动解压出来的 csv 文件到指定目录，并删除临时目录
            path_with_csv_file = unzip_target_path + temp_file
            target_path_with_file_name = unzip_target_path + file_name + '.csv'
            shutil.move(path_with_csv_file, target_path_with_file_name)
            if temp_path != None:
                temp_path = unzip_target_path + temp_path
            if util.is_directory_empty(temp_path):    
                os.rmdir(temp_path)
    except zipfile.BadZipFile:
        print("错误：文件不是一个有效的 zip 文件。")
    except RuntimeError as e:
        print(f"错误：{e}")
    except Exception as e:
        print(f"发生未知错误：{e}")

def handle_zfb_excel(path):
    file_path = path + 'Me-支付宝.csv'
    my_zfb_tx_list = parse_zfb_csv(file_path, "Me支付宝 ")
    file_path = path + 'Wife-支付宝.csv'
    wife_zfb_tx_list = parse_zfb_csv(file_path, "Wife支付宝 ")

    return sorted(my_zfb_tx_list + wife_zfb_tx_list, key=lambda tx: -tx.amount)

def handle_wx_excel(path):
    file_path = path + 'Me-微信.csv'
    my_wx_transaction_list = parse_wx_csv(file_path, 'MeWeChat ')
    file_path = path + 'Wife-微信.csv'
    wife_wx_transaction_list = parse_wx_csv(file_path, 'WifeWeChat ')
    
    transaction_list = merge_wx_tx(wife_wx_transaction_list, my_wx_transaction_list)
    return sorted(transaction_list, key=lambda tx: -tx.amount)

def merge_wx_tx(wife_wx_tx_list, my_wx_tx_list):
    # 合并两个交易列表，过滤掉我交易列表中的 "亲属卡交易"
    return wife_wx_tx_list + list(filter(lambda tx : tx.is_family_card() == False, my_wx_tx_list))

def xlsx_export(file_path, tx_data_array):
    workbook = openpyxl.Workbook()

    # 获取工作簿中的活动工作表（默认是第一个工作表）
    sheet = workbook.active

    # 写入表头
    sheet['A1'] = '日期'
    sheet['B1'] = '消费类型'
    sheet['C1'] = '消费方式'
    sheet['D1'] = '金额'
    sheet['E1'] = '备注'

    # 写入数据
    for row_idx, tx in enumerate(tx_data_array, start=2):  # 遍历数据，从第二行开始写入
        sheet.cell(row=row_idx, column=1, value=tx.date)
        sheet.cell(row=row_idx, column=2, value=tx.consume_type())
        sheet.cell(row=row_idx, column=3, value=tx.consume_way)
        sheet.cell(row=row_idx, column=4, value=tx.amount)
        sheet.cell(row=row_idx, column=5, value=tx.note)

    # 保存工作簿到文件
    workbook.save(file_path)

if __name__ == "__main__":
    main()
